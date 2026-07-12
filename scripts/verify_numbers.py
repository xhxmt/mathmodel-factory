#!/usr/bin/env python3
"""
verify_numbers.py - Two-way binding for numerical values in papers

Step 10 Gate 1: Generate numbers_manifest.json from results/*
Step 11+: Verify paper numbers against manifest with checksums

Usage:
  # Generate manifest (run at Step 10)
  python3 verify_numbers.py --generate <project_dir>

  # Verify paper against manifest (run at Step 11+)
  python3 verify_numbers.py --verify <project_dir> <base_name>

  # Update manifest after results change (Step 6+ reruns)
  python3 verify_numbers.py --update <project_dir>

  # Legacy mode (backward compatibility with old Step 10)
  python3 verify_numbers.py <project_dir> <base_name>
"""

import glob
import json
import os
import re
import sys
import hashlib
from pathlib import Path
from typing import Dict, List, Tuple, Any


def compute_checksum(value: Any) -> str:
    """Compute MD5 checksum for a value."""
    if isinstance(value, (int, float)):
        # Normalize float representation
        normalized = f"{value:.10e}" if isinstance(value, float) else str(value)
    else:
        normalized = str(value)
    return hashlib.md5(normalized.encode()).hexdigest()[:8]


def extract_tex_numbers(tex_path: str | Path) -> List[Dict[str, Any]]:
    """Extract prose numbers from a paper, matching the legacy helper behavior."""
    with open(tex_path, "r", encoding="utf-8", errors="replace") as f:
        text = f.read()

    doc_start = text.find(r"\begin{document}")
    if doc_start >= 0:
        text = text[doc_start:]

    text = re.sub(
        r'\\(?:label|ref|input|includegraphics|cite[tp]?|citealp|hypersetup|bibliographystyle|bibliography)\{[^}]*\}',
        '',
        text,
    )
    text = re.sub(r'\\(?:begin|end)\{[^}]*\}', '', text)
    text = re.sub(r'\\begin\{tabular\}.*?\\end\{tabular\}', '', text, flags=re.DOTALL)

    results: List[Dict[str, Any]] = []
    for line in text.splitlines():
        if line.strip().startswith('%'):
            continue
        if re.match(r'^\s*\\(setcounter|renewcommand|newcommand|def\\)', line):
            continue

        for match in re.finditer(r'-?\d[\d,]*\.?\d*%?', line):
            num_str = match.group()
            if re.match(r'^(19|20)\d{2}$', num_str):
                continue
            clean = num_str.replace(',', '').rstrip('%')
            try:
                value = float(clean)
            except ValueError:
                continue
            if value == 0:
                continue
            if value == int(value) and 1 <= value <= 20 and '.' not in num_str:
                continue

            start = max(0, match.start() - 40)
            end = min(len(line), match.end() + 40)
            context = line[start:end].strip()
            results.append({
                "number": num_str,
                "value": value,
                "is_pct": num_str.endswith('%'),
                "context": context,
            })
    return results


def extract_log_numbers(log_dir: str | Path) -> set[float]:
    numbers: set[float] = set()
    for lf in glob.glob(os.path.join(str(log_dir), "*.log")):
        try:
            with open(lf, "r", encoding="utf-8", errors="replace") as f:
                text = f.read()
        except OSError:
            continue

        for match in re.finditer(r'-?\d[\d,]*\.?\d*', text):
            clean = match.group().replace(',', '')
            try:
                value = float(clean)
            except ValueError:
                continue
            if value != 0:
                numbers.add(value)
    return numbers


def extract_table_numbers(tables_dir: str | Path) -> set[float]:
    numbers: set[float] = set()
    for tf in glob.glob(os.path.join(str(tables_dir), "*.tex")):
        try:
            with open(tf, "r", encoding="utf-8", errors="replace") as f:
                text = f.read()
        except OSError:
            continue

        for match in re.finditer(r'-?\d[\d,]*\.?\d*', text):
            clean = match.group().replace(',', '')
            try:
                value = float(clean)
            except ValueError:
                continue
            if value != 0:
                numbers.add(value)
    return numbers


def number_matches(val: float, reference_numbers: set[float], tolerance: float = 0.02) -> bool:
    if val in reference_numbers:
        return True
    for ref in reference_numbers:
        if ref == 0:
            continue
        if abs(val - ref) / abs(ref) <= tolerance:
            return True
        if abs(val - ref * 100) / abs(ref * 100) <= tolerance:
            return True
        if ref != 0 and abs(val - ref / 100) / abs(ref / 100) <= tolerance:
            return True
        if abs(val - ref * 1000) / abs(ref * 1000) <= tolerance:
            return True
    return False


def collect_number_metrics(project_dir: str | Path, base_name: str) -> Dict[str, Any] | None:
    """Backward-compatible numeric metrics API consumed by hard_metrics.py."""
    project_dir = Path(project_dir)
    tex_path = project_dir / f"{base_name}_paper.tex"
    if not tex_path.exists():
        return None

    log_dir = project_dir / "logs"
    tables_dir = project_dir / "tables"
    paper_numbers = extract_tex_numbers(tex_path)
    log_numbers = extract_log_numbers(log_dir)
    table_numbers = extract_table_numbers(tables_dir)
    reference_numbers = log_numbers | table_numbers
    matched = []
    unmatched = []
    for entry in paper_numbers:
        (matched if number_matches(entry["value"], reference_numbers) else unmatched).append(entry)

    return {
        "numbers_matched": len(matched),
        "numbers_unmatched": len(unmatched),
        "_matched": matched,
        "_unmatched": unmatched,
        "_paper_numbers": paper_numbers,
        "_reference_count": len(reference_numbers),
        "_log_count": len(log_numbers),
        "_table_count": len(table_numbers),
    }


def scan_results_directory(project_dir: Path) -> Dict[str, Any]:
    """
    Scan results/ directory and build manifest of all numerical values.

    Returns:
        {
            "source": "results/p1/values.json",
            "values": {
                "objective_value": {"value": 187.2, "checksum": "a3f5b2c1", "type": "float"},
                ...
            }
        }
    """
    manifest = {}
    results_dir = project_dir / "results"

    def add_numeric(source: str, key: str, value: int | float) -> None:
        if isinstance(value, bool):
            return
        manifest.setdefault(source, {})[key] = {
            "value": value,
            "checksum": compute_checksum(value),
            "type": "float" if isinstance(value, float) else "int",
        }

    def walk_json(value: Any, source: str, prefix: str = "") -> None:
        if isinstance(value, bool):
            return
        if isinstance(value, (int, float)):
            if prefix:
                add_numeric(source, prefix, value)
            return
        if isinstance(value, dict):
            for child_key, child_value in value.items():
                child = f"{prefix}.{child_key}" if prefix else str(child_key)
                walk_json(child_value, source, child)
        elif isinstance(value, list):
            for idx, child_value in enumerate(value):
                child = f"{prefix}[{idx}]" if prefix else f"[{idx}]"
                walk_json(child_value, source, child)

    if results_dir.exists():
        for json_file in sorted(results_dir.rglob("*.json")):
            try:
                with open(json_file, encoding="utf-8") as f:
                    data = json.load(f)
                relative_path = str(json_file.relative_to(project_dir))
                walk_json(data, relative_path)
            except Exception as e:
                print(f"Warning: Failed to parse {json_file}: {e}", file=sys.stderr)

    try:
        from openpyxl import load_workbook
    except ImportError:
        load_workbook = None

    if load_workbook is not None:
        for xlsx_file in sorted(project_dir.glob("result*.xlsx")):
            try:
                wb = load_workbook(xlsx_file, data_only=True, read_only=True)
                relative_path = str(xlsx_file.relative_to(project_dir))
                for ws in wb.worksheets:
                    for row in ws.iter_rows():
                        for cell in row:
                            value = cell.value
                            if isinstance(value, bool):
                                continue
                            if isinstance(value, (int, float)):
                                add_numeric(relative_path, f"{ws.title}!{cell.coordinate}", value)
                wb.close()
            except Exception as e:
                print(f"Warning: Failed to parse {xlsx_file}: {e}", file=sys.stderr)

    return manifest


def generate_manifest(project_dir: Path) -> None:
    """Generate numbers_manifest.json from results/*"""
    manifest = scan_results_directory(project_dir)

    output = {
        "generated_by": "scripts/verify_numbers.py",
        "step": "Step 10 Gate 1",
        "sources": manifest
    }

    manifest_file = project_dir / "numbers_manifest.json"
    with open(manifest_file, "w") as f:
        json.dump(output, f, indent=2)

    print(f"✓ Generated {manifest_file}")
    print(f"  Total sources: {len(manifest)}")
    print(f"  Total values: {sum(len(v) for v in manifest.values())}")


def _strip_non_content_latex(line: str) -> str:
    """Remove LaTeX commands whose numeric arguments are formatting, not results."""
    line = re.sub(r'\\includegraphics(?:\[[^\]]*\])?\{[^}]*\}', '', line)
    line = re.sub(
        r'\\(?:label|ref|pageref|cite[tp]?|citealp|hypersetup|bibliographystyle|bibliography)\{[^}]*\}',
        '',
        line,
    )
    line = re.sub(r'\\(?:begin|end)\{[^}]*\}', '', line)
    # Avoid false positives from LaTeX command names and exponent notation:
    # \approx1.23 should expose 1.23, not a bogus 23; 10^{-6} should not
    # emit -6 as a standalone result number.
    line = re.sub(r'\^\{[-+]?\d+\}', '', line)
    line = re.sub(r'\^[+-]?\d+', '', line)
    line = re.sub(r'\\[a-zA-Z]+\*?', ' ', line)
    line = line.replace(r'\,', ' ')
    return line


def extract_numbers_from_tex(tex_file: Path) -> List[Tuple[int, str, float]]:
    """
    Extract numerical values from LaTeX file.

    Returns:
        [(line_number, context, value), ...]
    """
    numbers = []

    in_document = False
    in_references = False

    with open(tex_file) as f:
        for line_num, line in enumerate(f, start=1):
            if r"\begin{document}" in line:
                in_document = True
                line = line.split(r"\begin{document}", 1)[1]
            if not in_document:
                continue
            if r"\section{参考文献}" in line or r"\begin{thebibliography}" in line:
                in_references = True
            if r"\appendix" in line:
                in_references = False
                continue
            if in_references:
                continue

            # Skip comments
            if line.strip().startswith('%'):
                continue

            line_for_numbers = _strip_non_content_latex(line)

            # Find numbers in text (not in commands)
            # Match patterns like: 187.2, 0.0478, $x = 42$, etc.
            pattern = r'(?<![a-zA-Z])-?\d+\.?\d*(?:[eE][+-]?\d+)?%?'
            for match in re.finditer(pattern, line_for_numbers):
                try:
                    raw = match.group(0)
                    clean = raw.rstrip('%')
                    value = float(clean)
                    if (
                        value == int(value)
                        and "." not in clean
                        and "e" not in clean.lower()
                        and (1 <= value <= 20 or 1900 <= value <= 2099 or value == 100)
                    ):
                        continue
                    context = line.strip()[:60]
                    numbers.append((line_num, context, value))
                except ValueError:
                    pass

    return numbers


def verify_paper(project_dir: Path, base_name: str) -> bool:
    """
    Verify paper numbers against manifest.

    Returns:
        True if all checks pass, False otherwise
    """
    manifest_file = project_dir / "numbers_manifest.json"
    if not manifest_file.exists():
        print(f"✗ numbers_manifest.json not found. Run with --generate first.", file=sys.stderr)
        return False

    with open(manifest_file) as f:
        manifest_data = json.load(f)

    # Verify that the manifest still reflects current result artifacts.
    current_sources = scan_results_directory(project_dir)
    source_mismatches = []

    # Build reverse lookup entries.  Do not collapse duplicate values: common
    # integers such as 0/1/2 may legitimately appear in multiple result files
    # with different JSON/XLSX scalar types.
    value_sources = []
    for source_path, values in manifest_data["sources"].items():
        for key, entry in values.items():
            value = entry["value"]
            checksum = entry["checksum"]
            if isinstance(value, list):
                value = tuple(value)
            current_entry = current_sources.get(source_path, {}).get(key)
            if current_entry is None:
                source_mismatches.append((source_path, key, "missing"))
            elif current_entry.get("checksum") != checksum:
                source_mismatches.append((source_path, key, "checksum"))
            value_sources.append((value, source_path, key, checksum))

    # Extract numbers from paper
    paper_file = project_dir / f"{base_name}_paper.tex"
    if not paper_file.exists():
        print(f"✗ Paper file {paper_file} not found.", file=sys.stderr)
        return False

    paper_numbers = extract_numbers_from_tex(paper_file)

    # Check each number
    untraced = []
    def matches(manifest_value: float, paper_value: float) -> bool:
        abs_tol = 1e-6
        rel_tol = 5e-3
        return abs(float(manifest_value) - paper_value) <= max(abs_tol, abs(float(manifest_value)) * rel_tol)

    for line_num, context, value in paper_numbers:
        # Try to find this value in manifest (with tolerance for floats)
        found = False
        for manifest_value, source, key, checksum in value_sources:
            if isinstance(manifest_value, (int, float)):
                if matches(float(manifest_value), value):
                    found = True
                    break

        if not found:
            untraced.append((line_num, context, value))

    # Report results
    report_file = project_dir / "number_verification.md"
    with open(report_file, "w") as f:
        f.write(f"# Number Verification Report — {base_name}\n\n")
        f.write(f"Paper: `{base_name}_paper.tex`\n")
        f.write(f"Manifest: `numbers_manifest.json`\n\n")

        if not untraced and not source_mismatches:
            f.write("✓ All numerical values are traced to source files.\n")
            print("✓ Number verification passed")
            return True

        if untraced:
            f.write(f"## ⚠ Untraced Numbers ({len(untraced)})\n\n")
            f.write("These numbers appear in the paper but not in `results/*/values.json`:\n\n")
            for line_num, context, value in untraced:
                f.write(f"- Line {line_num}: `{value}` in \"{context}\"\n")
            f.write("\n")

        if source_mismatches:
            f.write(f"## ⚠ Source Checksum Mismatches ({len(source_mismatches)})\n\n")
            f.write("The manifest is stale relative to current result artifacts:\n\n")
            for source, key, reason in source_mismatches:
                f.write(f"- `{source}::{key}` ({reason})\n")
            f.write("\n")

    print(f"✗ Number verification found issues. See {report_file}")
    return False


def update_manifest(project_dir: Path) -> None:
    """Update manifest after results/ changes."""
    print("Updating numbers_manifest.json...")
    generate_manifest(project_dir)


def _legacy_verify(project_dir: Path, base_name: str) -> int:
    """Preserve the old positional CLI output used by existing tests/tools."""
    tex_path = project_dir / f"{base_name}_paper.tex"
    if not tex_path.exists():
        print(f"ERROR: {tex_path} not found")
        return 1

    display_tex_path = str(tex_path.resolve())
    worktree_marker = "/.worktrees/"
    if worktree_marker in display_tex_path:
        _, after = display_tex_path.split(worktree_marker, 1)
        parts = after.split("/", 1)
        if len(parts) == 2:
            repo_root = tex_path.resolve().parents[5]
            display_tex_path = str(repo_root / parts[1])

    metrics = collect_number_metrics(project_dir, base_name)
    matched = metrics["_matched"]
    unmatched = metrics["_unmatched"]

    print("=== Number Verification Report ===")
    print(f"Paper: {display_tex_path}")
    print(f"Log numbers found: {metrics['_log_count']}")
    print(f"Table numbers found: {metrics['_table_count']}")
    print(f"Paper prose numbers found: {len(metrics['_paper_numbers'])}")
    print()
    print(f"MATCHED (found in logs/tables): {len(matched)}")
    print(f"UNMATCHED (no source found):    {len(unmatched)}")
    print()

    if unmatched:
        print("=" * 70)
        print("UNMATCHED NUMBERS — review these manually:")
        print("=" * 70)
        for entry in unmatched:
            print(f"  {entry['number']:>12s}  ...{entry['context']}...")
        print()

    if matched:
        print("=" * 70)
        print("MATCHED NUMBERS (OK):")
        print("=" * 70)
        for entry in matched:
            print(f"  {entry['number']:>12s}  ...{entry['context']}...")

    return 0 if not unmatched else 1


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    # Legacy mode: verify_numbers.py <project_dir> <base_name>
    if len(sys.argv) == 3 and not sys.argv[1].startswith('--'):
        project_dir = Path(sys.argv[1])
        base_name = sys.argv[2]
        sys.exit(_legacy_verify(project_dir, base_name))

    # New mode with flags
    mode = sys.argv[1]
    if len(sys.argv) < 3:
        print("Error: Missing project_dir argument", file=sys.stderr)
        print(__doc__)
        sys.exit(1)

    project_dir = Path(sys.argv[2])

    if not project_dir.exists():
        print(f"Error: Project directory {project_dir} not found", file=sys.stderr)
        sys.exit(1)

    if mode == "--generate":
        generate_manifest(project_dir)
    elif mode == "--verify":
        if len(sys.argv) < 4:
            print("Error: --verify requires base_name argument", file=sys.stderr)
            sys.exit(1)
        base_name = sys.argv[3]
        success = verify_paper(project_dir, base_name)
        sys.exit(0 if success else 1)
    elif mode == "--update":
        update_manifest(project_dir)
    else:
        print(f"Error: Unknown mode {mode}", file=sys.stderr)
        print(__doc__)
        sys.exit(1)


if __name__ == "__main__":
    main()
