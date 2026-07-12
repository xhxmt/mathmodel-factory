import json
import os
import shutil
import subprocess
import sys
import zipfile
from pathlib import Path

from conftest import REPO_ROOT


def write_file(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def write_zip(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w") as zf:
        zf.writestr("paper.pdf", "fake")


def make_project_with_delivery(factory: Path, base: str, verdict: str = "PASS") -> Path:
    project = factory / "ongoing" / base
    paper = project / f"{base}_paper.tex"
    write_file(project / "problem" / "problem_brief.md", "# brief\n")
    write_file(project / "checkpoint.md", "- **Last completed step**: 15\n")
    write_file(project / "judge_evaluation.md", f"VERDICT: {verdict}\n" + "\n".join(["judge"] * 30) + "\n")
    write_file(project / "reviewer_entry_map.md", "# map\n")
    write_file(project / "anchor_figure_plan.md", "# anchors\n")
    write_file(project / "entry_gate.md", "VERDICT: PASS\n")
    write_file(paper, "\\begin{document}\n" + "\n".join(["paper"] * 220) + "\n\\end{document}\n")
    write_file(factory / "papers" / f"{base}_paper.pdf", "%PDF fake\n")
    write_zip(factory / "papers" / f"{base}_submission.zip")
    return project


def test_infer_step_does_not_report_16_when_gate2_is_not_pass(tmp_path):
    project = make_project_with_delivery(tmp_path, "demo_reopen", "REOPEN_REVISION_TEXT")

    out = subprocess.run(
        [os.path.join(REPO_ROOT, "run_paper.sh"), "--infer-step", str(project)],
        env={**os.environ, "FACTORY": str(tmp_path)},
        capture_output=True,
        text=True,
        check=False,
    )

    assert out.returncode == 0, out.stderr
    assert out.stdout.strip() != "16"


def test_step16_writes_delivery_manifest_after_quality_gate():
    runner = Path(REPO_ROOT) / "run_paper.sh"
    text = runner.read_text(encoding="utf-8")

    assert "scripts/delivery_contract.py" in text
    assert "delivery_manifest.json" in text
    assert text.index("Final delivery quality gate PASS") < text.index("delivery_manifest.json")


def test_verify_step_output_rejects_step9_without_step8_5_pass(tmp_path):
    project = tmp_path / "ongoing" / "demo_step9"
    write_file(project / "problem" / "problem_brief.md", "# brief\n")
    write_file(project / "checkpoint.md", "- **Last completed step**: 8\n")
    write_file(project / "visualization_log.md", "\n".join(["viz"] * 20) + "\n")
    write_file(project / "figures" / "anchor.pdf", "fake\n")
    write_file(
        project / "demo_step9_paper.tex",
        "\\begin{document}\nABSTRACT_PLACEHOLDER\n" + "\n".join(["paper"] * 220) + "\n\\end{document}\n",
    )

    out = subprocess.run(
        [os.path.join(REPO_ROOT, "run_paper.sh"), "--infer-step", str(project)],
        env={**os.environ, "FACTORY": str(tmp_path)},
        capture_output=True,
        text=True,
        check=False,
    )

    assert out.returncode == 0, out.stderr
    assert out.stdout.strip() == "8"


def test_evaluator_rejects_incomplete_canonical_results(tmp_path, monkeypatch):
    project = tmp_path / "ongoing" / "demo"
    from test_evaluate_modeling_project_step8_5 import make_complete_project

    make_complete_project(project)
    write_file(project / "reviewer_entry_map.md", "# map\n")
    write_file(project / "anchor_figure_plan.md", "# anchors\n")
    write_file(project / "entry_gate.md", "VERDICT: PASS\n")
    write_file(project / "results" / "p1" / "values.json", '{"problem": 1, "status": "CONVERGED", "objective": 1.0}\n')
    write_file(project / "results" / "p2" / "values.json", '{"problem": 2, "status": "RUNNING"}\n')

    from scripts import evaluate_modeling_project as mod

    monkeypatch.setattr(mod, "infer_step", lambda root, project: (16, "16"))
    monkeypatch.setattr(mod, "zip_ok", lambda path: (True, "ok"))

    ev = mod.evaluate(project, tmp_path)
    checks = {check.name: check for check in ev.checks}

    assert checks["canonical_results"].ok is False
    assert "p2" in checks["canonical_results"].detail


def test_evaluator_and_delivery_contract_reject_failed_reality_gates(tmp_path, monkeypatch):
    project = tmp_path / "ongoing" / "demo_reality_gate"
    from test_evaluate_modeling_project_step8_5 import make_complete_project

    make_complete_project(project)
    write_file(tmp_path / "method_library" / "demo.md", "# demo method\n")
    write_file(project / "reviewer_entry_map.md", "# map\n")
    write_file(project / "anchor_figure_plan.md", "# anchors\n")
    write_file(project / "entry_gate.md", "VERDICT: PASS\n")
    write_file(project / "numbers_manifest.json", '{"sources": {}}\n')
    write_file(project / "results" / "canonical_results.json", '{"objective": 1.0, "status": "FEASIBLE"}\n')

    from scripts import delivery_contract
    from scripts import evaluate_modeling_project as mod

    monkeypatch.setattr(mod, "infer_step", lambda root, project: (16, "16"))
    monkeypatch.setattr(mod, "zip_ok", lambda path: (True, "ok"))
    monkeypatch.setattr(mod, "symbol_check_ok", lambda root, project, base: (True, "verify_symbols PASS"))

    def fake_run_python_check(root, args, timeout=60):
        script = args[0]
        if script == "scripts/verify_provenance.py":
            return False, "VERDICT: REPAIR_FALLBACK"
        if script == "scripts/verify_spec_impl.py":
            return False, "VERDICT: FAIL"
        return True, "VERDICT: PASS"

    monkeypatch.setattr(mod, "run_python_check", fake_run_python_check)

    ev = mod.evaluate(project, tmp_path)
    checks = {check.name: check for check in ev.checks}

    assert checks["provenance_gate"].ok is False
    assert checks["spec_impl_gate"].ok is False
    assert ev.passed is False

    manifest = delivery_contract.build_delivery_manifest(project, tmp_path, ev, generated_at="2026-07-10T00:00:00+00:00")

    assert manifest["status"] != "CURRENT_PASS"
    assert {check["name"] for check in manifest["evaluation"]["failed_checks"]} >= {
        "provenance_gate",
        "spec_impl_gate",
    }


def test_evaluator_rejects_failed_project_quality_contract(tmp_path, monkeypatch):
    project = tmp_path / "ongoing" / "demo_quality_contract"
    from test_evaluate_modeling_project_step8_5 import make_complete_project

    make_complete_project(project)
    write_file(tmp_path / "method_library" / "demo.md", "# demo method\n")
    write_file(project / "reviewer_entry_map.md", "# map\n")
    write_file(project / "anchor_figure_plan.md", "# anchors\n")
    write_file(project / "entry_gate.md", "VERDICT: PASS\n")
    write_file(project / "numbers_manifest.json", '{"sources": {}}\n')
    write_file(project / "results" / "canonical_results.json", '{"objective": 1.0, "status": "FEASIBLE"}\n')
    write_file(project / "quality_contract.json", '{"version": 1, "claims": [], "anomaly_checks": []}\n')

    from scripts import evaluate_modeling_project as mod

    monkeypatch.setattr(mod, "infer_step", lambda root, project: (16, "16"))
    monkeypatch.setattr(mod, "zip_ok", lambda path: (True, "ok"))
    monkeypatch.setattr(mod, "symbol_check_ok", lambda root, project, base: (True, "verify_symbols PASS"))

    def fake_run_python_check(root, args, timeout=60):
        if args[0] == "scripts/verify_quality_contract.py":
            return False, "QUALITY_CONTRACT_FAILURES=1\nVERDICT: FAIL"
        return True, "VERDICT: PASS"

    monkeypatch.setattr(mod, "run_python_check", fake_run_python_check)

    ev = mod.evaluate(project, tmp_path)
    checks = {check.name: check for check in ev.checks}

    assert checks["quality_contract_gate"].ok is False
    assert ev.passed is False


def test_runner_requires_quality_contract_at_gate1_and_delivery():
    runner = (Path(REPO_ROOT) / "run_paper.sh").read_text(encoding="utf-8")

    assert "quality_contract_gate_passed()" in runner
    verifier = runner[runner.index("verify_step_output()") :]
    step10_start = verifier.index("        10)")
    step10 = verifier[step10_start : verifier.index("        11)", step10_start)]
    delivery = runner[
        runner.index("step16_hard_acceptance()") : runner.index("infer_step()")
    ]

    assert 'quality_contract_gate_passed "$P"' in step10
    assert 'quality_contract_gate_passed "$P"' in delivery


def test_step4_is_not_complete_without_quality_contract(tmp_path):
    project = tmp_path / "ongoing" / "demo_step4_contract"
    write_file(project / "problem" / "problem_brief.md", "# brief\n")
    write_file(project / "method_decision.md", "\n".join(["decision"] * 35) + "\n")
    write_file(project / "chosen_method.md", "PRIMARY: m1 family=test\n" + "\n".join(["chosen"] * 10) + "\n")
    write_file(project / "model.md", "\n".join(["model"] * 110) + "\n")
    write_file(project / "symbol_table.md", "\n".join(["symbol"] * 12) + "\n")
    write_file(project / "assumption_ledger.md", "\n".join(["assumption"] * 12) + "\n")
    write_file(project / "modeling_scope_gate.md", "VERDICT: PASS\n")

    out = subprocess.run(
        [os.path.join(REPO_ROOT, "run_paper.sh"), "--infer-step", str(project)],
        env={**os.environ, "FACTORY": str(tmp_path)},
        capture_output=True,
        text=True,
        check=False,
    )

    assert out.returncode == 0, out.stderr
    assert out.stdout.strip() == "3"


def test_verify_numbers_manifest_includes_nested_json_and_xlsx(tmp_path):
    project = tmp_path / "proj"
    write_file(project / "results" / "p1" / "values.json", json.dumps({
        "problem": 1,
        "objective": 4.9,
        "decision": {"theta_deg": 8.7, "v_mps": 140.0},
        "intervals": [[1.5, 6.4]],
    }))
    try:
        from openpyxl import Workbook
    except ImportError:
        return
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "duration"
    ws["B1"] = 14.804
    wb.save(project / "result3.xlsx")

    from verify_numbers import scan_results_directory

    manifest = scan_results_directory(project)
    flat_keys = set(manifest["results/p1/values.json"].keys())

    assert "decision.theta_deg" in flat_keys
    assert "intervals[0][1]" in flat_keys
    assert manifest["result3.xlsx"]["Sheet!B1"]["value"] == 14.804


def test_verify_numbers_rejects_unrecorded_simple_derived_number(tmp_path):
    project = tmp_path / "proj"
    base = "proj"
    write_file(project / "results" / "p1" / "values.json", json.dumps({"a": 2.0, "b": 3.0}))
    write_file(project / f"{base}_paper.tex", "\\begin{document}\nThe reported value is 5.0.\n\\end{document}\n")

    from verify_numbers import generate_manifest, verify_paper

    generate_manifest(project)

    assert verify_paper(project, base) is False
    report = (project / "number_verification.md").read_text(encoding="utf-8")
    assert "5.0" in report


def test_verify_numbers_handles_latex_commands_and_exponents(tmp_path):
    project = tmp_path / "proj"
    base = "proj"
    write_file(project / "results" / "p1" / "values.json", json.dumps({"mask_time_s": 1.3624}))
    write_file(
        project / f"{base}_paper.tex",
        "\\begin{document}\n"
        "The verified result is $T_1\\approx1.3624\\,\\mathrm{s}$ with tolerance $<10^{-6}$.\n"
        "\\end{document}\n",
    )

    from verify_numbers import generate_manifest, verify_paper

    generate_manifest(project)

    assert verify_paper(project, base) is True


def test_verify_symbols_treats_big_set_operators_as_latex_noise(tmp_path):
    project = tmp_path / "proj"
    base = "proj"
    write_file(project / "symbol_table.md", "| 符号 | 含义 |\n|---|---|\n| $I$ | interval |\n| $j$ | index |\n")
    write_file(
        project / f"{base}_paper.tex",
        "\\begin{document}\n"
        "\\section{符号说明}\n"
        "Registered variables are $I$ and $j$.\n"
        "The aggregate is $\\bigcup_{j=1}^{3} I_j$ and $\\bigcap_{j=1}^{3} I_j$.\n"
        "\\end{document}\n",
    )

    from verify_symbols import collect_symbol_metrics

    metrics = collect_symbol_metrics(project, base)

    assert "\\bigcup" not in metrics["_undefined_list"]
    assert "\\bigcap" not in metrics["_undefined_list"]


def test_verify_symbols_ignores_norm_delimiters_and_equation_labels(tmp_path):
    project = tmp_path / "proj"
    base = "proj"
    write_file(
        project / "symbol_table.md",
        "| 符号 | 含义 |\n|---|---|\n| $x$ | vector |\n| $t$ | time |\n",
    )
    write_file(
        project / f"{base}_paper.tex",
        "\\begin{document}\n"
        "\\section{符号说明}\n"
        "Registered variables are $x$ and $t$.\n"
        "\\begin{equation}\n"
        "\\lVert x\\rVert = t.\n"
        "\\label{eq:active_window}\n"
        "\\end{equation}\n"
        "\\end{document}\n",
    )

    from verify_symbols import collect_symbol_metrics

    metrics = collect_symbol_metrics(project, base)

    assert metrics["_undefined_list"] == []


def test_runner_invokes_step3_selection_before_step3_dispatch():
    runner = Path(REPO_ROOT) / "run_paper.sh"
    text = runner.read_text(encoding="utf-8")

    assert "maybe_select_option step3 3" in text
    assert text.index("maybe_select_option step3 3") < text.index("3)  run_step_3")


def test_runner_retry_branch_uses_global_variables_not_local_declarations():
    runner = Path(REPO_ROOT) / "run_paper.sh"
    text = runner.read_text(encoding="utf-8")

    retry_branch = text[text.index("RETRIES=$((RETRIES + 1))") : text.index("if (( RETRIES >= STEP_MAX_RETRIES ))")]

    assert "\n            local " not in retry_branch
    assert "\n                local " not in retry_branch


def test_dispatch_step_validates_artifacts_after_successful_configured_backend():
    runner = Path(REPO_ROOT) / "run_paper.sh"
    text = runner.read_text(encoding="utf-8")
    dispatch = text[text.index("dispatch_step() {") : text.index("# Wrapper so single-Claude-worker steps")]

    success_branch = dispatch[dispatch.index('if run_backend "$primary"') : dispatch.index('if [[ -n "$fallback"')]

    verifier = 'if verify_step "$NEXT" && verify_step_output "$NEXT"; then'
    assert verifier in success_branch
    assert success_branch.index(verifier) < success_branch.index("return 0")


def test_codex_only_mode_blocks_claude_backend_and_default_fallbacks():
    runner = Path(REPO_ROOT) / "run_paper.sh"
    text = runner.read_text(encoding="utf-8")
    backend = text[text.index("run_backend() {") : text.index("# Generic per-step dispatch")]
    dispatch = text[text.index("dispatch_step() {") : text.index("# Wrapper so single-Claude-worker steps")]
    wrappers = text[text.index("run_agy_then_claude() {") : text.index("run_step_1()")]

    assert "codex_only_enabled()" in text
    assert 'if codex_only_enabled; then' in backend
    assert 'Claude backend blocked because CODEX_ONLY=1' in backend
    assert 'CODEX_ONLY=1: using Codex instead of configured/default model dispatch' in dispatch
    assert 'CODEX_ONLY=1: using Codex instead of built-in default' in dispatch
    assert 'CODEX_ONLY=1: configured model(s) failed; not invoking built-in default' in dispatch
    assert 'CODEX_ONLY=1: Claude fallback disabled' in wrappers


def test_claude_worker_has_stale_activity_kill_path():
    runner = Path(REPO_ROOT) / "run_paper.sh"
    text = runner.read_text(encoding="utf-8")
    worker = text[text.index("run_claude_worker() {") : text.index("# Fallback: Codex failed")]

    assert 'hang_timeout="${5:-3600}"' in worker
    assert ") > \"$claude_log\" 2>&1 &" in worker
    assert "local claude_pid=$!" in worker
    assert 'while kill -0 "$claude_pid"' in worker
    assert "now - stale_since > hang_timeout" in worker
    assert '_kill_process_tree "$claude_pid" TERM' in worker
    assert '_kill_process_tree "$claude_pid" KILL' in worker


def test_claude_backend_passes_dispatch_hang_timeout_to_worker():
    runner = Path(REPO_ROOT) / "run_paper.sh"
    text = runner.read_text(encoding="utf-8")
    backend = text[text.index("run_backend() {") : text.index("# Generic per-step dispatch")]
    wrapper = text[text.index("_default_claude_worker()") : text.index("run_step_1()")]

    assert 'run_claude_worker "$prompt_file" "$timeout" "" "$model" "$hang"' in backend
    assert '_default_claude_worker() { run_claude_worker "$1" "$2" "" "" "$3"; }' in wrapper


def test_step6_precheck_parses_markdown_assumption_table(tmp_path):
    project = tmp_path / "proj"
    write_file(project / "assumption_ledger.md", "\n".join([
        "| id | 陈述 | 来源 | 若违反的影响 | 状态 | 标签 |",
        "|---|---|---|---|---|---|",
        "| A1 | assumption | source | impact | INHERITED | **PROTECTED** |",
        "| A2 | assumption | source | impact | OPEN | CRITICAL |",
    ]))

    from scripts.step6_coverage_precheck import check_assumption_ledger

    ok, msg, counts = check_assumption_ledger(project)

    assert ok is True
    assert counts["INHERITED"] == 1
    assert counts["OPEN"] == 1


def test_project_monitor_once_handles_missing_results_dirs():
    project_name = "monitor_missing_results_fixture"
    project = Path(REPO_ROOT) / "ongoing" / project_name
    monitor_log = Path(REPO_ROOT) / "run_state" / f"{project_name}_monitor.log"

    shutil.rmtree(project, ignore_errors=True)
    monitor_log.unlink(missing_ok=True)
    write_file(project / "checkpoint.md", "- **Last completed step**: 0\n")
    write_file(project / "problem" / "problem_brief.md", "# brief\n")
    write_file(project / "logs" / "runner.log", "runner started\n")

    try:
        out = subprocess.run(
            [os.path.join(REPO_ROOT, "scripts", "project_monitor.sh"), "--once", project_name],
            cwd=REPO_ROOT,
            capture_output=True,
            text=True,
            check=False,
        )

        assert out.returncode == 0, out.stderr
        report = monitor_log.read_text(encoding="utf-8")
        assert "values_json_count=0" in report
        assert "solver_job_meta_count=0" in report
    finally:
        shutil.rmtree(project, ignore_errors=True)
        monitor_log.unlink(missing_ok=True)
